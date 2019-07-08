#excel文档操作

import os
import json,re
import openpyxl
import shutil
from openpyxl.styles import *
from openpyxl.styles.colors import RED
import openpyxl.styles as sty

#获取目标文件夹的路径
fileDir = 'D:\\modefiles'
#上报属性和模型里属性一致的文件的路径
noProDir = 'D:\\noHandle'
#上报属性和模型里属性不一致，文件处理后的路径
proDir = 'D:\\handled'
#没有查到上报日志的路径
noLogPath = 'D:\\inexisLog'

noHandle = []
handled = []

#fileName = '非标准模型对应关系表-插座-201c80c70c50031c14060dcfe615544ee1689867960aa2e4976054b446e38640-1406.xlsx'
fileName = ''
attrs = ''''''
colsIndex = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
fileLogName = 'typeidAttrsLog.xlsx'
typeidAttrLog = []
typeidAttrLog2 = []
attrLog = []

def read_excel():
    # 打开文件 fieleHandles,attrsJson
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.worksheets[0]
    # for i in sheet[1]:
    #     print(i.value) #获取第1行的值
    #获取第2行数据
    row2Value = tuple2list(sheet,2,None)
    #print(row2Value)
    insert_clo(row2Value,workbook,sheet)


    #print(sheet.max_row) #获取最多行
    #cell = sheet['B1']
    #print(cell.row, cell.column)#获取单元格对应的行号和列值

    colsList,colsListType,mergedAttsCols = get_merged_cells_value(sheet, 0, 0)
    #print('*********** ' )
    #print(colsList)
    #print(colsListType)
    #colsListOnly = get_only_list(colsList)
    #colsListTypeOnly = get_only_list(colsListType)
    #print(colsListOnly)
    #print(colsListTypeOnly)




    #attrs = '''{"attrs":[{"name":"onOffStatus","value":"1"},{"name":"position","value":"7"}],"edata":null}'''


    jsonAttrs = json.loads(attrs,strict=False)
    attrsList = jsonAttrs['attrs']

    attrsListName = get_attrs_name(attrsList)
    attrsListValue = get_attrs_vlue(attrsList)

    # 判断exle里的name是否在上报的状态中
    flag1 = False
    for i in range(len(colsList)):
        if (i >= 2) and (colsList[i] not in attrsListName):
            print('不在attrs里：')
            #print(colsList[i])
            operate_excel2(workbook, sheet, i + 1, colsIndex.index(mergedAttsCols) - 1, colsList[i],True)
            flag1 = True
    #if flag1:
        #剪切到
        # 剪切/移动
        #shutil.move(fileDir + '\\' + fileName, proDir)





    #判断上报状态的name是否在excel里
    flag2 = False
    flag3 = False
    for name in attrsListName:
        #print("name= ",name)
        if name in colsList:
            #print("True")
            # 获取上报状态的value
            attrValueIndex = attrsListName.index(name)
            attrValue = attrsListValue[attrValueIndex]

            for i in range(len(colsList)):
                if i >= 2:
                    if colsList[i] == name:
                        #判断是否是bool
                        if 'bool' == colsListType[i]:
                            # 判断value是否属于bool类型
                            m = re.search(attrValue, 'true false', re.IGNORECASE)
                            if m:
                                print('++++++++++++++++' + attrValue)
                                flag3 = True
                                operate_excel(workbook, sheet, i + 1, colsIndex.index(mergedAttsCols), False, attrValue)
                            else:
                                print('-----------------')
                                flag2 = True
                                operate_excel(workbook,sheet,i+1,colsIndex.index(mergedAttsCols),True,attrValue)
                        else:
                            flag3 = True
                            print('++++++++++++++++'+attrValue)
                            # 写到exce里
                            operate_excel(workbook, sheet, i + 1, colsIndex.index(mergedAttsCols), False, attrValue)
        else:
            # name没有在excel里
            print(' name没有在excel里')
            flag2 = True
            operate_excel2(workbook, sheet, sheet.max_row+1 , colsIndex.index(mergedAttsCols)-1, name,False)

    if flag1 or flag2:
        # 剪切/移动
        shutil.move( fileName, proDir)
    if flag3 :
        # 剪切/移动
        shutil.move(fileName, noProDir)



def operate_excel(workbook,sheet,index,cols,fontFlag,value):
    font = Font(color=RED)
    for i ,row in enumerate(sheet.rows):
        if i == index-1:
            if fontFlag:
                row[cols+1].font = font
                row[cols+1].value = value
            else:
                row[cols+1].value = value

    workbook.save(fileName)
def operate_excel2(workbook,sheet,index,cols,value,flag):
    font = Font(color=RED)
    fill = sty.PatternFill(fill_type="solid", fgColor="FFF68F")
    sheet.cell(index,cols).font = font
    if flag:
        sheet.cell(index, cols).fill = fill
    sheet.cell(index, cols).value = value
    workbook.save(fileName)


def tuple2list(sheet,rows,cols):
    valueList = []
    #len(sheet[rows]) sheet.max_row
    if cols is None:
        for i in range(len(sheet[rows]) ):
            valueList.append(sheet.cell(rows, i + 1).value)
    else:
        #print('有'+str(sheet.max_row)+'行')
        for i in range(sheet.max_row):
            #print('********cols '+cols)
            #print('********' + str(colsIndex.index(cols) + 1))
            valueList.append(sheet.cell(i + 1, colsIndex.index(cols) + 1).value)


    return valueList

def insert_clo(row2List,workbook,sheet):
    index = 0
    fill = sty.PatternFill(fill_type="solid", fgColor="FFF68F")
    for i in range(len(row2List)):
        if '枚举值描述' == row2List[i]:
            index = i
    sheet.insert_cols(index+1)#插入一列
    for i, row in enumerate(sheet.rows):
        row[index].fill = fill
        if i == 1:
            row[index ].value = u'上报值'
    workbook.save(fileName)


def get_attrs_name(listAttr):
    listArrName = []
    for i in iter(listAttr):
        for key in i:
            if key == 'name':
                name = (i[key])
                listArrName.append(name)
    return listArrName

def get_attrs_vlue(listAttr):
    listArrValue = []
    for i in iter(listAttr):
        for key in i:
            if key == 'value':
                name = (i[key])
                listArrValue.append(name)
    return listArrValue


#去重列表重复元素
def get_only_list(repeatList):
    list2 = []
    list1 = repeatList
    for i in list1:
        if i not in list2:
            list2.append(i)
    return list2

def get_merged_cells(sheet):
    #print('合并的单元格：')
    #print(sheet.merged_cells)
    # 合并单元格的起始行坐标、终止行坐标
    return sheet.merged_cells


def get_merged_cells_value(sheet, row_index, col_index):
    """
    先判断给定的单元格，是否属于合并单元格；
    如果是合并单元格，就返回合并单元格的内容
    :return:
    """
    merged = get_merged_cells(sheet)
    mergedList = str(merged).split(':')
    colNum = mergedList[1][3] #获取第二个合并单元格的开始列
    #print(colNum)
    #print(chr(ord(colNum)+1))
    #print(sheet[chr(ord(colNum)+1)])# 取某一列的数据
    #print('!!!!!&&&&& ' + colNum)
    nameList = tuple2list(sheet,1,chr(ord(colNum)+1))
    typeList = tuple2list(sheet,1,chr(ord(colNum)+2))
    attrCol = chr(ord(colNum)+3)
    return (nameList,typeList,attrCol)



def read_log_excel():
    # 打开文件
    workbook = openpyxl.load_workbook(fileLogName)
    sheet = workbook.worksheets[0]

    #获取第一列typeid并去重
    for i, row in enumerate(sheet.rows):
        if i >= 1:
            if row[0].value not in typeidAttrLog:
                typeidAttrLog.append(row[0].value)
                #print(row[0].value)
    #print(len(typeidAttrLog))


    for i in range(len(typeidAttrLog)):
        log = typeidAttrLog[i]
        for j, row in enumerate(sheet.rows):
            j = j+1
            if (log == sheet['A' + str(j)].value) and ('devrpt' == sheet['U' + str(j)].value) and ('''{"attrs"''' == sheet['AC' + str(j)].value[:8]):
                attrLog.append(sheet['AC' + str(j)].value)
                typeidAttrLog2.append(log)
                #print(log)
                #print(sheet['AC' + str(j)].value)
                #print(j)
                break


def go_split(s, symbol):
    # 拼接正则表达式
    symbol = "[" + symbol + "]+"
    # 一次性分割字符串
    result = re.split(symbol, s)
    # 去除空字符
    return [x for x in result if x]

#获取指定目录下的源文件
def get_files(path):
    return os.listdir(path)

#获取文件名里的typeid
def get_typeid(fileNames):
    typeids = []
    for dir in fileNames:
        typeid = go_split(dir, '-')
        typeids.append(typeid[2])

    return typeids





if __name__ == "__main__":
    #read_excel()
    #读取日志文件获取上报的状态属性和typeid
    read_log_excel()
    #print(len(typeidAttrLog2))
    #print(len(attrLog))
    # 获取待处理指定目录下的文件
    fileNames = get_files(fileDir)
    #k开始处理
    for file in fileNames:
        fileType = go_split(file, '-')
        fileTypeId = fileType[2]
        #print(fileTypeId)
        #获取attr日志
        attrIndex = typeidAttrLog2.index(fileTypeId)
        attrlog = attrLog[attrIndex]
        attrs = attrlog
        fileName = fileDir + '\\'+file
        print(file)
        #read_excel()






    #剪切日志里不存在的typeid文件
    '''
    fileNamess = get_files(noLogPath)
    for typeid in typeidAttrLog2:

        for file in fileNamess:
            print(file)
            if typeid in file:
                # 剪切/移动
                shutil.move(noLogPath + '\\' + file, fileDir)
               
    print('ok')
    '''









